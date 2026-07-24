"""v54: пілот транспортних листів для одного водія.

v57: транспортні листи увімкнено для всіх активних водіїв із закріпленим авто.
Пробіг для списання пального — різниця одометра; GPS лише довідковий.
"""
import io
import os
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

router = APIRouter(tags=["fuel"])
pool = None
KYIV_TZ = ZoneInfo("Europe/Kyiv")


async def init(db_pool):
    global pool
    pool = db_pool
    await pool.execute("""
        CREATE TABLE IF NOT EXISTS vehicle_fuel_settings (
            vehicle_id INT PRIMARY KEY REFERENCES vehicles(id) ON DELETE CASCADE,
            rate_l_per_100 NUMERIC(8,3), initial_balance_l NUMERIC(10,3),
            initial_balance_date DATE, updated_at TIMESTAMPTZ NOT NULL DEFAULT now())""")
    await pool.execute("""
        CREATE TABLE IF NOT EXISTS transport_sheets (
            id BIGSERIAL PRIMARY KEY, work_date DATE NOT NULL,
            vehicle_id INT NOT NULL REFERENCES vehicles(id),
            driver_id INT NOT NULL REFERENCES drivers(id),
            odometer_start NUMERIC(12,1), odometer_end NUMERIC(12,1),
            opening_balance_l NUMERIC(10,3), fuel_used_l NUMERIC(10,3),
            closing_balance_l NUMERIC(10,3),
            status TEXT NOT NULL DEFAULT 'draft'
                CHECK (status IN ('draft','submitted','approved','revision')),
            revision_reason TEXT, submitted_at TIMESTAMPTZ, approved_at TIMESTAMPTZ,
            approved_by TEXT, created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            UNIQUE (work_date, vehicle_id))""")
    await pool.execute("CREATE INDEX IF NOT EXISTS idx_transport_sheets_date ON transport_sheets(work_date DESC)")
    await pool.execute("CREATE INDEX IF NOT EXISTS idx_transport_sheets_driver ON transport_sheets(driver_id, work_date DESC)")
    await pool.execute("""
        CREATE TABLE IF NOT EXISTS transport_sheet_refuels (
            id BIGSERIAL PRIMARY KEY, sheet_id BIGINT NOT NULL REFERENCES transport_sheets(id) ON DELETE CASCADE,
            liters NUMERIC(10,3) NOT NULL CHECK (liters > 0),
            refuel_at TIMESTAMPTZ NOT NULL DEFAULT now(), created_at TIMESTAMPTZ NOT NULL DEFAULT now())""")
    await pool.execute("""
        CREATE TABLE IF NOT EXISTS transport_sheet_changes (
            id BIGSERIAL PRIMARY KEY, sheet_id BIGINT NOT NULL REFERENCES transport_sheets(id) ON DELETE CASCADE,
            actor_role TEXT NOT NULL CHECK (actor_role IN ('driver','logist','system')),
            actor_name TEXT, field_name TEXT NOT NULL, old_value TEXT, new_value TEXT,
            reason TEXT, created_at TIMESTAMPTZ NOT NULL DEFAULT now())""")
    # v55
    await pool.execute(
        "ALTER TABLE transport_sheets ADD COLUMN IF NOT EXISTS odometer_start_confirmed_at TIMESTAMPTZ")
    await pool.execute(
        "ALTER TABLE route_events ADD COLUMN IF NOT EXISTS source TEXT NOT NULL DEFAULT 'driver'")
    await pool.execute("""
        CREATE TABLE IF NOT EXISTS fuel_balance_adjustments (
            id BIGSERIAL PRIMARY KEY,
            vehicle_id INT NOT NULL REFERENCES vehicles(id) ON DELETE CASCADE,
            adjust_date DATE NOT NULL,
            balance_l NUMERIC(10,3) NOT NULL CHECK (balance_l >= 0),
            reason TEXT NOT NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now())""")
    await pool.execute("""CREATE INDEX IF NOT EXISTS idx_fuel_adjustments_vehicle
        ON fuel_balance_adjustments(vehicle_id, adjust_date DESC, id DESC)""")


async def _derive_opening(vehicle_id, day, conn=None):
    """v55: залишок на початок дня — ТІЛЬКИ з ланцюжка.

    Пріоритет: (а) коригування, датоване пізніше за останній підтверджений день;
    (б) залишок кін. останнього підтвердженого дня; (в) стартова точка з налаштувань.
    Жодних фолбеків на одометр чи ручні значення.
    """
    c = conn or pool
    prev = await c.fetchrow("""
        SELECT work_date, closing_balance_l FROM transport_sheets
        WHERE vehicle_id=$1 AND work_date<$2 AND status='approved'
          AND closing_balance_l IS NOT NULL
        ORDER BY work_date DESC LIMIT 1""", vehicle_id, day)
    adj = await c.fetchrow("""
        SELECT adjust_date, balance_l FROM fuel_balance_adjustments
        WHERE vehicle_id=$1 AND adjust_date<=$2
        ORDER BY adjust_date DESC, id DESC LIMIT 1""", vehicle_id, day)
    if adj and (not prev or adj["adjust_date"] > prev["work_date"]):
        return adj["balance_l"]
    if prev:
        return prev["closing_balance_l"]
    settings = await c.fetchrow(
        "SELECT initial_balance_l, initial_balance_date FROM vehicle_fuel_settings WHERE vehicle_id=$1",
        vehicle_id)
    if settings and settings["initial_balance_l"] is not None and \
       settings["initial_balance_date"] and settings["initial_balance_date"] <= day:
        return settings["initial_balance_l"]
    return None


def _num(value, field, allow_none=True):
    if value is None and allow_none:
        return None
    try:
        result = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise HTTPException(400, f"{field}: невірне число")
    if result < 0:
        raise HTTPException(400, f"{field}: значення не може бути від'ємним")
    return result


async def _driver(token):
    row = await pool.fetchrow("""
        SELECT d.id, d.name, d.code_1c FROM driver_tokens t
        JOIN drivers d ON d.id=t.driver_id
        WHERE t.token=$1 AND t.is_active AND d.is_active""", token)
    if not row:
        raise HTTPException(401, "Недійсний токен")
    return row


async def _vehicle_for(driver_id, day):
    row = await pool.fetchrow("""
        SELECT v.id, v.name, v.plate FROM routes r JOIN vehicles v ON v.id=r.vehicle_id
        WHERE r.plan_date=$2 AND (r.driver_id=$1 OR v.driver_id=$1)
        ORDER BY r.depart_time NULLS LAST, r.id LIMIT 1""", driver_id, day)
    if not row:
        row = await pool.fetchrow(
            "SELECT id,name,plate FROM vehicles WHERE driver_id=$1 AND is_active ORDER BY id LIMIT 1",
            driver_id)
    return row


async def _ensure_sheet(driver, day):
    vehicle = await _vehicle_for(driver["id"], day)
    if not vehicle:
        return None, None
    sheet = await pool.fetchrow(
        "SELECT * FROM transport_sheets WHERE work_date=$1 AND vehicle_id=$2", day, vehicle["id"])
    if sheet:
        return sheet, vehicle
    previous = await pool.fetchrow("""
        SELECT odometer_end FROM transport_sheets
        WHERE vehicle_id=$1 AND work_date<$2 AND odometer_end IS NOT NULL AND status='approved'
        ORDER BY work_date DESC LIMIT 1""", vehicle["id"], day)
    opening = await _derive_opening(vehicle["id"], day)          # v55: тільки ланцюжок
    start = previous["odometer_end"] if previous else None
    sheet = await pool.fetchrow("""
        INSERT INTO transport_sheets (work_date,vehicle_id,driver_id,odometer_start,opening_balance_l)
        VALUES ($1,$2,$3,$4,$5)
        ON CONFLICT (work_date,vehicle_id) DO UPDATE SET driver_id=EXCLUDED.driver_id
        RETURNING *""", day, vehicle["id"], driver["id"], start, opening)
    return sheet, vehicle


async def _recalculate(sheet_id):
    row = await pool.fetchrow("""
        SELECT s.*, f.rate_l_per_100,
               COALESCE((SELECT sum(liters) FROM transport_sheet_refuels x WHERE x.sheet_id=s.id),0) refueled_l
        FROM transport_sheets s LEFT JOIN vehicle_fuel_settings f ON f.vehicle_id=s.vehicle_id
        WHERE s.id=$1""", sheet_id)
    if not row:
        raise HTTPException(404, "Лист не знайдено")
    km = None
    if row["odometer_start"] is not None and row["odometer_end"] is not None:
        km = row["odometer_end"] - row["odometer_start"]
        if km < 0:
            raise HTTPException(400, "Кінцевий одометр не може бути меншим за початковий")
    opening = row["opening_balance_l"]
    if row["status"] != "approved":                              # v68: завжди звіряти з ланцюжком
        opening = await _derive_opening(row["vehicle_id"], row["work_date"])
        if opening != row["opening_balance_l"]:
            await pool.execute("""INSERT INTO transport_sheet_changes
                (sheet_id,actor_role,actor_name,field_name,old_value,new_value,reason)
                VALUES ($1,'system','Система','opening_balance_l',$2,$3,
                        'Автоматичне відновлення паливного ланцюжка')""",
                sheet_id,
                str(row["opening_balance_l"]) if row["opening_balance_l"] is not None else None,
                str(opening) if opening is not None else None)
    used = km * row["rate_l_per_100"] / 100 if km is not None and row["rate_l_per_100"] is not None else None
    closing = (opening + row["refueled_l"] - used
               if opening is not None and used is not None else None)
    if (opening == row["opening_balance_l"]
            and used == row["fuel_used_l"]
            and closing == row["closing_balance_l"]):
        return row
    return await pool.fetchrow("""
        UPDATE transport_sheets SET opening_balance_l=$1,fuel_used_l=$2,closing_balance_l=$3,updated_at=now()
        WHERE id=$4 RETURNING *""", opening, used, closing, sheet_id)


async def _payload(sheet, vehicle):
    sheet = await _recalculate(sheet["id"])
    cfg = await pool.fetchrow("SELECT * FROM vehicle_fuel_settings WHERE vehicle_id=$1", sheet["vehicle_id"])
    refs = await pool.fetch("SELECT id,liters,refuel_at FROM transport_sheet_refuels WHERE sheet_id=$1 ORDER BY refuel_at,id", sheet["id"])
    changes = await pool.fetch("""
        SELECT actor_role,actor_name,field_name,old_value,new_value,reason,created_at
        FROM transport_sheet_changes WHERE sheet_id=$1 ORDER BY created_at DESC LIMIT 30""", sheet["id"])
    km = (sheet["odometer_end"] - sheet["odometer_start"]
          if sheet["odometer_start"] is not None and sheet["odometer_end"] is not None else None)
    return {"id": sheet["id"], "date": sheet["work_date"].isoformat(),
            "vehicle_id": sheet["vehicle_id"], "vehicle": vehicle["name"], "plate": vehicle["plate"],
            "driver_id": sheet["driver_id"], "odometer_start": sheet["odometer_start"],
            "odometer_end": sheet["odometer_end"], "km": km,
            "opening_balance_l": sheet["opening_balance_l"], "refueled_l": sum(r["liters"] for r in refs),
            "fuel_used_l": sheet["fuel_used_l"], "closing_balance_l": sheet["closing_balance_l"],
            "rate_l_per_100": cfg["rate_l_per_100"] if cfg else None,
            "status": sheet["status"], "revision_reason": sheet["revision_reason"],
            "start_confirmed": sheet["odometer_start_confirmed_at"] is not None,   # v55
            "settings_ready": bool(cfg and cfg["rate_l_per_100"] is not None),
            "refuels": [dict(r) for r in refs], "changes": [dict(c) for c in changes]}


@router.get("/api/driver/{token}/transport-sheet")
async def driver_sheet(token: str, d: date | None = Query(None)):
    driver = await _driver(token)
    day = d or datetime.now(KYIV_TZ).date()
    sheet, vehicle = await _ensure_sheet(driver, day)
    if not sheet:
        return {"enabled": False}          # v57: без авто — інтерфейс ТЛ не показуємо
    return {"enabled": True, "sheet": await _payload(sheet, vehicle)}


class OdometerIn(BaseModel):
    odometer_start: str | float | None = None
    odometer_end: str | float | None = None
    reason: str | None = None
    confirm_start: bool = False          # v55: ранкове підтвердження одним тапом


async def _pilot_sheet(token, day=None):
    driver = await _driver(token)
    sheet, vehicle = await _ensure_sheet(driver, day or datetime.now(KYIV_TZ).date())
    if not sheet:
        raise HTTPException(400, "За водієм не закріплено автомобіль")
    return driver, sheet, vehicle


@router.post("/api/driver/{token}/transport-sheet/odometer")
async def save_odometer(token: str, body: OdometerIn):
    driver, sheet, vehicle = await _pilot_sheet(token)
    start, end = _num(body.odometer_start, "Початковий одометр"), _num(body.odometer_end, "Кінцевий одометр")
    if start is None and end is None:
        raise HTTPException(400, "Вкажіть показник")
    if sheet["status"] == "approved" and not (body.reason or "").strip():
        raise HTTPException(400, "Після підтвердження обов'язково вкажіть причину")
    new_start = start if start is not None else sheet["odometer_start"]
    new_end = end if end is not None else sheet["odometer_end"]
    if new_start is not None and new_end is not None and new_end < new_start:
        raise HTTPException(400, "Кінцевий одометр не може бути меншим за початковий")
    reason = (body.reason or "").strip() or None
    async with pool.acquire() as c:
        async with c.transaction():
            for field, old, new in (("odometer_start", sheet["odometer_start"], new_start),
                                    ("odometer_end", sheet["odometer_end"], new_end)):
                if new != old:
                    await c.execute("""INSERT INTO transport_sheet_changes
                        (sheet_id,actor_role,actor_name,field_name,old_value,new_value,reason)
                        VALUES ($1,'driver',$2,$3,$4,$5,$6)""",
                        sheet["id"], driver["name"], field,
                        str(old) if old is not None else None, str(new), reason)
            status = "revision" if sheet["status"] == "approved" else "draft"
            confirm = body.confirm_start and new_start is not None
            sheet = await c.fetchrow("""UPDATE transport_sheets SET odometer_start=$1,odometer_end=$2,
                status=$3,revision_reason=$4,approved_at=NULL,approved_by=NULL,
                odometer_start_confirmed_at=CASE WHEN $6 THEN now()
                    ELSE odometer_start_confirmed_at END,updated_at=now()
                WHERE id=$5 RETURNING *""", new_start, new_end, status, reason, sheet["id"], confirm)
    return {"ok": True, "sheet": await _payload(sheet, vehicle)}


class RefuelIn(BaseModel):
    liters: str | float
    reason: str | None = None


@router.post("/api/driver/{token}/transport-sheet/refuels")
async def add_refuel(token: str, body: RefuelIn):
    driver, sheet, vehicle = await _pilot_sheet(token)
    liters = _num(body.liters, "Літри", False)
    if liters <= 0:
        raise HTTPException(400, "Кількість літрів має бути більшою за нуль")
    if sheet["status"] == "approved" and not (body.reason or "").strip():
        raise HTTPException(400, "Після підтвердження обов'язково вкажіть причину")
    reason = (body.reason or "").strip() or None
    async with pool.acquire() as c:
        async with c.transaction():
            ref_id = await c.fetchval(
                "INSERT INTO transport_sheet_refuels(sheet_id,liters) VALUES($1,$2) RETURNING id",
                sheet["id"], liters)
            await c.execute("""INSERT INTO transport_sheet_changes
                (sheet_id,actor_role,actor_name,field_name,new_value,reason)
                VALUES ($1,'driver',$2,'refuel',$3,$4)""", sheet["id"], driver["name"], str(liters), reason)
            if sheet["status"] == "approved":
                await c.execute("""UPDATE transport_sheets SET status='revision',revision_reason=$1,
                    approved_at=NULL,approved_by=NULL,updated_at=now() WHERE id=$2""", reason, sheet["id"])
    return {"ok": True, "refuel_id": ref_id,
            "sheet": await _payload(await pool.fetchrow("SELECT * FROM transport_sheets WHERE id=$1", sheet["id"]), vehicle)}


@router.post("/api/driver/{token}/transport-sheet/submit")
async def submit_sheet(token: str):
    driver, sheet, vehicle = await _pilot_sheet(token)
    if sheet["odometer_start"] is None or sheet["odometer_end"] is None:
        raise HTTPException(400, "Внесіть початковий і кінцевий одометр")
    cfg = await pool.fetchrow("SELECT * FROM vehicle_fuel_settings WHERE vehicle_id=$1", sheet["vehicle_id"])
    if not cfg or cfg["rate_l_per_100"] is None or sheet["opening_balance_l"] is None:
        raise HTTPException(400, "Логіст ще не налаштував норму та початковий залишок")
    await _recalculate(sheet["id"])
    sheet = await pool.fetchrow("""UPDATE transport_sheets SET status='submitted',submitted_at=now(),
        revision_reason=NULL,updated_at=now() WHERE id=$1 RETURNING *""", sheet["id"])
    return {"ok": True, "sheet": await _payload(sheet, vehicle)}


class SettingsIn(BaseModel):
    rate_l_per_100: str | float
    initial_balance_l: str | float | None = None       # v55: опційно, керується коригуваннями
    initial_balance_date: date | None = None


@router.put("/api/transport-sheets/settings/{vehicle_id}")
async def save_settings(vehicle_id: int, body: SettingsIn):
    rate = _num(body.rate_l_per_100, "Норма", False)
    if rate <= 0:
        raise HTTPException(400, "Норма має бути більшою за нуль")
    existing = await pool.fetchrow(
        "SELECT initial_balance_l, initial_balance_date FROM vehicle_fuel_settings WHERE vehicle_id=$1",
        vehicle_id)
    balance = _num(body.initial_balance_l, "Початковий залишок") \
        if body.initial_balance_l is not None else (existing["initial_balance_l"] if existing else None)
    balance_date = body.initial_balance_date or (existing["initial_balance_date"] if existing else None)
    await pool.execute("""INSERT INTO vehicle_fuel_settings
        (vehicle_id,rate_l_per_100,initial_balance_l,initial_balance_date)
        VALUES ($1,$2,$3,$4) ON CONFLICT(vehicle_id) DO UPDATE SET
        rate_l_per_100=EXCLUDED.rate_l_per_100,initial_balance_l=EXCLUDED.initial_balance_l,
        initial_balance_date=EXCLUDED.initial_balance_date,updated_at=now()""",
        vehicle_id, rate, balance, balance_date)
    # v55: перерахувати всі непідтверджені листи від дати стартової точки
    await _recalc_chain(vehicle_id, balance_date or datetime.now(KYIV_TZ).date(),
                        "Оновлення налаштувань")
    updated = await pool.fetchval(
        "SELECT updated_at FROM vehicle_fuel_settings WHERE vehicle_id=$1", vehicle_id)
    return {"ok": True, "updated_at": updated.isoformat() if updated else None}


# ---------- v55: датовані коригування залишку ----------

async def _recalc_chain(vehicle_id: int, from_date: date, reason: str):
    """Перерахувати ланцюжок залишків уперед від from_date.

    Підтверджені листи не чіпаємо (їх цифри вже пішли в облік) — вони
    залишаються опорними точками; перераховуються draft/submitted/revision.
    """
    sheets = await pool.fetch("""
        SELECT id, work_date, status, opening_balance_l FROM transport_sheets
        WHERE vehicle_id=$1 AND work_date>=$2 ORDER BY work_date""", vehicle_id, from_date)
    for s in sheets:
        if s["status"] == "approved":
            continue
        opening = await _derive_opening(vehicle_id, s["work_date"])
        if opening != s["opening_balance_l"]:
            await pool.execute("""INSERT INTO transport_sheet_changes
                (sheet_id,actor_role,actor_name,field_name,old_value,new_value,reason)
                VALUES ($1,'system','Система','opening_balance_l',$2,$3,$4)""",
                s["id"],
                str(s["opening_balance_l"]) if s["opening_balance_l"] is not None else None,
                str(opening) if opening is not None else None, reason)
        await pool.execute(
            "UPDATE transport_sheets SET opening_balance_l=$1,updated_at=now() WHERE id=$2",
            opening, s["id"])
        await _recalculate(s["id"])


class AdjustmentIn(BaseModel):
    vehicle_id: int
    adjust_date: date
    balance_l: str | float
    reason: str


@router.post("/api/transport-sheets/adjustments")
async def add_adjustment(body: AdjustmentIn):
    balance = _num(body.balance_l, "Залишок", False)
    reason = (body.reason or "").strip()
    if not reason:
        raise HTTPException(400, "Вкажіть причину коригування")
    blocking = await pool.fetchrow("""
        SELECT work_date FROM transport_sheets
        WHERE vehicle_id=$1 AND work_date>=$2 AND status='approved'
        ORDER BY work_date DESC LIMIT 1""", body.vehicle_id, body.adjust_date)
    if blocking:
        raise HTTPException(
            409,
            f"Спочатку скасуйте підтвердження за {blocking['work_date'].isoformat()}")
    await pool.execute("""INSERT INTO fuel_balance_adjustments
        (vehicle_id,adjust_date,balance_l,reason) VALUES ($1,$2,$3,$4)""",
        body.vehicle_id, body.adjust_date, balance, reason)
    await _recalc_chain(body.vehicle_id, body.adjust_date, f"Коригування залишку: {reason}")
    return {"ok": True}


@router.get("/api/transport-sheets/balance/{vehicle_id}")
async def fuel_balance(vehicle_id: int):
    last = await pool.fetchrow("""
        SELECT work_date, closing_balance_l, status FROM transport_sheets
        WHERE vehicle_id=$1 AND closing_balance_l IS NOT NULL
        ORDER BY work_date DESC LIMIT 1""", vehicle_id)
    settings = await pool.fetchrow(
        "SELECT initial_balance_l, initial_balance_date FROM vehicle_fuel_settings WHERE vehicle_id=$1",
        vehicle_id)
    adjustments = await pool.fetch("""
        SELECT id, adjust_date, balance_l, reason, created_at
        FROM fuel_balance_adjustments WHERE vehicle_id=$1
        ORDER BY adjust_date DESC, id DESC LIMIT 20""", vehicle_id)
    return {"current": dict(last) if last else None,
            "start_point": dict(settings) if settings else None,
            "adjustments": [dict(a) for a in adjustments]}


def _sheet_dict(r):
    return {k: r[k] for k in r.keys()}


@router.get("/api/transport-sheets")
async def list_sheets(date_from: date, date_to: date, driver_ids: str | None = None):
    ids = [int(x) for x in (driver_ids or "").split(",") if x.strip().isdigit()]
    # v68: самовідновлення відкритих листів після коригування або скасування підтвердження.
    open_rows = await pool.fetch("""
        SELECT id FROM transport_sheets
        WHERE work_date BETWEEN $1 AND $2 AND status<>'approved'
          AND ($3::int[] IS NULL OR driver_id=ANY($3))""",
        date_from, date_to, ids or None)
    for row in open_rows:
        await _recalculate(row["id"])
    rows = await pool.fetch("""
        SELECT s.*,d.name driver_name,d.code_1c,v.name vehicle_name,v.plate,
               f.rate_l_per_100,COALESCE(sum(rf.liters),0) refueled_l
        FROM transport_sheets s JOIN drivers d ON d.id=s.driver_id
        JOIN vehicles v ON v.id=s.vehicle_id
        LEFT JOIN vehicle_fuel_settings f ON f.vehicle_id=s.vehicle_id
        LEFT JOIN transport_sheet_refuels rf ON rf.sheet_id=s.id
        WHERE s.work_date BETWEEN $1 AND $2 AND ($3::int[] IS NULL OR s.driver_id=ANY($3))
        GROUP BY s.id,d.name,d.code_1c,v.name,v.plate,f.rate_l_per_100
        ORDER BY s.work_date DESC,d.name""", date_from, date_to, ids or None)
    return [_sheet_dict(r) for r in rows]


@router.get("/api/transport-sheets/meta")
async def sheet_meta():
    rows = await pool.fetch("""
        SELECT d.id driver_id,d.name,d.code_1c,v.id vehicle_id,v.name vehicle_name,v.plate,
               f.rate_l_per_100,f.initial_balance_l,f.initial_balance_date
        FROM drivers d LEFT JOIN vehicles v ON v.driver_id=d.id AND v.is_active
        LEFT JOIN vehicle_fuel_settings f ON f.vehicle_id=v.id
        WHERE d.is_active AND v.id IS NOT NULL ORDER BY d.name""")
    return {"drivers": [dict(r) for r in rows]}


class LogistSheetIn(BaseModel):
    odometer_start: str | float | None = None
    odometer_end: str | float | None = None
    opening_balance_l: str | float | None = None
    reason: str


@router.patch("/api/transport-sheets/{sheet_id}")
async def edit_sheet(sheet_id: int, body: LogistSheetIn):
    sheet = await pool.fetchrow("SELECT * FROM transport_sheets WHERE id=$1", sheet_id)
    if not sheet:
        raise HTTPException(404, "Лист не знайдено")
    if sheet["status"] == "approved":
        raise HTTPException(409, "Спочатку скасуйте підтвердження листа")
    reason = (body.reason or "").strip()
    if not reason:
        raise HTTPException(400, "Вкажіть причину зміни")
    # v55: opening_balance_l ігнорується — ним керує ланцюжок і коригування
    vals = {"odometer_start": _num(body.odometer_start, "Початок"),
            "odometer_end": _num(body.odometer_end, "Кінець")}
    if vals["odometer_start"] is not None and vals["odometer_end"] is not None and vals["odometer_end"] < vals["odometer_start"]:
        raise HTTPException(400, "Кінцевий одометр не може бути меншим")
    async with pool.acquire() as c:
        async with c.transaction():
            for field,new in vals.items():
                if new != sheet[field]:
                    await c.execute("""INSERT INTO transport_sheet_changes
                        (sheet_id,actor_role,actor_name,field_name,old_value,new_value,reason)
                        VALUES ($1,'logist','Логіст',$2,$3,$4,$5)""", sheet_id, field,
                        str(sheet[field]) if sheet[field] is not None else None,
                        str(new) if new is not None else None, reason)
            await c.execute("""UPDATE transport_sheets SET odometer_start=$1,odometer_end=$2,
                status='submitted',updated_at=now() WHERE id=$3""",
                vals["odometer_start"], vals["odometer_end"], sheet_id)
    await _recalculate(sheet_id)
    return {"ok": True}


# ---------- v56: логіст редагує заправки ----------

class RefuelEditIn(BaseModel):
    liters: str | float
    reason: str


async def _touch_after_refuel_edit(sheet):
    """Після зміни заправки: перерахунок; підтверджений лист повертається на перевірку."""
    if sheet["status"] == "approved":
        await pool.execute("""UPDATE transport_sheets SET status='submitted',
            approved_at=NULL,approved_by=NULL,updated_at=now() WHERE id=$1""", sheet["id"])
    await _recalculate(sheet["id"])


@router.patch("/api/transport-sheets/{sheet_id}/refuels/{refuel_id}")
async def logist_edit_refuel(sheet_id: int, refuel_id: int, body: RefuelEditIn):
    sheet = await pool.fetchrow("SELECT * FROM transport_sheets WHERE id=$1", sheet_id)
    if not sheet:
        raise HTTPException(404, "Лист не знайдено")
    if sheet["status"] == "approved":
        raise HTTPException(409, "Спочатку скасуйте підтвердження листа")
    ref = await pool.fetchrow(
        "SELECT * FROM transport_sheet_refuels WHERE id=$1 AND sheet_id=$2", refuel_id, sheet_id)
    if not ref:
        raise HTTPException(404, "Заправку не знайдено")
    liters = _num(body.liters, "Літри", False)
    if liters <= 0:
        raise HTTPException(400, "Кількість літрів має бути більшою за нуль")
    reason = (body.reason or "").strip()
    if not reason:
        raise HTTPException(400, "Вкажіть причину зміни")
    if liters == ref["liters"]:
        return {"ok": True}
    async with pool.acquire() as c:
        async with c.transaction():
            await c.execute(
                "UPDATE transport_sheet_refuels SET liters=$1 WHERE id=$2", liters, refuel_id)
            await c.execute("""INSERT INTO transport_sheet_changes
                (sheet_id,actor_role,actor_name,field_name,old_value,new_value,reason)
                VALUES ($1,'logist','Логіст','refuel',$2,$3,$4)""",
                sheet_id, str(ref["liters"]), str(liters), reason)
    await _touch_after_refuel_edit(sheet)
    return {"ok": True}


@router.delete("/api/transport-sheets/{sheet_id}/refuels/{refuel_id}")
async def logist_delete_refuel(sheet_id: int, refuel_id: int,
                               reason: str = Query(..., min_length=1)):
    sheet = await pool.fetchrow("SELECT * FROM transport_sheets WHERE id=$1", sheet_id)
    if not sheet:
        raise HTTPException(404, "Лист не знайдено")
    if sheet["status"] == "approved":
        raise HTTPException(409, "Спочатку скасуйте підтвердження листа")
    ref = await pool.fetchrow(
        "SELECT * FROM transport_sheet_refuels WHERE id=$1 AND sheet_id=$2", refuel_id, sheet_id)
    if not ref:
        raise HTTPException(404, "Заправку не знайдено")
    async with pool.acquire() as c:
        async with c.transaction():
            await c.execute("DELETE FROM transport_sheet_refuels WHERE id=$1", refuel_id)
            await c.execute("""INSERT INTO transport_sheet_changes
                (sheet_id,actor_role,actor_name,field_name,old_value,new_value,reason)
                VALUES ($1,'logist','Логіст','refuel',$2,NULL,$3)""",
                sheet_id, str(ref["liters"]), reason.strip())
    await _touch_after_refuel_edit(sheet)
    return {"ok": True}


@router.post("/api/transport-sheets/{sheet_id}/reopen")
async def reopen_sheet(sheet_id: int):
    """v67: логіст скасовує останнє підтвердження без втрати введених даних."""
    async with pool.acquire() as c:
        async with c.transaction():
            sheet = await c.fetchrow(
                "SELECT * FROM transport_sheets WHERE id=$1 FOR UPDATE", sheet_id)
            if not sheet:
                raise HTTPException(404, "Лист не знайдено")
            if sheet["status"] != "approved":
                raise HTTPException(409, "Лист уже не підтверджений")
            later = await c.fetchrow("""
                SELECT work_date FROM transport_sheets
                WHERE vehicle_id=$1 AND work_date>$2 AND status='approved'
                ORDER BY work_date DESC LIMIT 1""",
                sheet["vehicle_id"], sheet["work_date"])
            if later:
                raise HTTPException(
                    409,
                    f"Спочатку скасуйте підтвердження за {later['work_date'].isoformat()}")
            await c.execute("""INSERT INTO transport_sheet_changes
                (sheet_id,actor_role,actor_name,field_name,old_value,new_value)
                VALUES ($1,'logist','Логіст','status','approved','submitted')""", sheet_id)
            await c.execute("""UPDATE transport_sheets SET status='submitted',
                approved_at=NULL,approved_by=NULL,updated_at=now()
                WHERE id=$1""", sheet_id)
    await _recalc_chain(
        sheet["vehicle_id"], sheet["work_date"], "Скасування підтвердження")
    return {"ok": True}


@router.post("/api/transport-sheets/{sheet_id}/approve")
async def approve_sheet(sheet_id: int):
    sheet = await pool.fetchrow("SELECT * FROM transport_sheets WHERE id=$1", sheet_id)
    if not sheet:
        raise HTTPException(404, "Лист не знайдено")
    if sheet["odometer_start"] is None or sheet["odometer_end"] is None:
        raise HTTPException(400, "Немає повних показників одометра")
    sheet = await _recalculate(sheet_id)
    if sheet["closing_balance_l"] is None:
        raise HTTPException(400, "Немає норми або початкового залишку")
    await pool.execute("""UPDATE transport_sheets SET status='approved',approved_at=now(),
        approved_by='Логіст',revision_reason=NULL,updated_at=now() WHERE id=$1""", sheet_id)
    return {"ok": True}


@router.get("/api/transport-sheets/notifications/count")
async def notification_count():
    count = await pool.fetchval("SELECT count(*) FROM transport_sheets WHERE status='revision'")
    return {"count": count}


@router.get("/api/transport-sheets/export.xlsx")
async def export_sheets(date_from: date, date_to: date, driver_ids: str | None = None,
                        approved_only: bool = True):
    ids = [int(x) for x in (driver_ids or "").split(",") if x.strip().isdigit()]
    rows = await pool.fetch("""
        SELECT s.*,d.name driver_name,v.name vehicle_name,v.plate,f.rate_l_per_100,
               COALESCE(sum(rf.liters),0) refueled_l
        FROM transport_sheets s JOIN drivers d ON d.id=s.driver_id
        JOIN vehicles v ON v.id=s.vehicle_id LEFT JOIN vehicle_fuel_settings f ON f.vehicle_id=s.vehicle_id
        LEFT JOIN transport_sheet_refuels rf ON rf.sheet_id=s.id
        WHERE s.work_date BETWEEN $1 AND $2
          AND ($3::int[] IS NULL OR s.driver_id=ANY($3))
          AND (NOT $4 OR s.status='approved')
        GROUP BY s.id,d.name,v.name,v.plate,f.rate_l_per_100
        ORDER BY v.name,s.work_date""", date_from, date_to, ids or None, approved_only)
    wb = Workbook(); ws = wb.active; ws.title = "Транспортні листи"
    headers = ["Дата","Автомобіль","Водій","Одометр початок","Одометр кінець","Пробіг, км",
               "Надходження, л","Норма, л/100 км","Витрата, л","Залишок початок, л",
               "Залишок кінець, л","Статус"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="00356B");cell.alignment=Alignment(horizontal="center")
    for r in rows:
        km = r["odometer_end"]-r["odometer_start"] if r["odometer_start"] is not None and r["odometer_end"] is not None else None
        ws.append([r["work_date"],r["vehicle_name"],r["driver_name"],r["odometer_start"],r["odometer_end"],km,
                   r["refueled_l"],r["rate_l_per_100"],r["fuel_used_l"],r["opening_balance_l"],r["closing_balance_l"],r["status"]])
    for width,col in zip([13,24,22,18,18,14,17,18,14,21,20,16],ws.columns):
        ws.column_dimensions[col[0].column_letter].width=width
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    name=f"transport_sheets_{date_from}_{date_to}.xlsx"
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":f'attachment; filename="{name}"'})


@router.get("/api/transport-sheets/{sheet_id}")
async def sheet_detail(sheet_id: int):
    await _recalculate(sheet_id)                    # v68: актуальний залишок перед відкриттям
    row = await pool.fetchrow("""
        SELECT s.*,d.name driver_name,v.name vehicle_name,v.plate,
               f.rate_l_per_100,COALESCE(sum(rf.liters),0) refueled_l
        FROM transport_sheets s JOIN drivers d ON d.id=s.driver_id
        JOIN vehicles v ON v.id=s.vehicle_id
        LEFT JOIN vehicle_fuel_settings f ON f.vehicle_id=s.vehicle_id
        LEFT JOIN transport_sheet_refuels rf ON rf.sheet_id=s.id
        WHERE s.id=$1 GROUP BY s.id,d.name,v.name,v.plate,f.rate_l_per_100""", sheet_id)
    if not row:
        raise HTTPException(404, "Лист не знайдено")
    changes = await pool.fetch("""SELECT actor_role,actor_name,field_name,old_value,new_value,reason,created_at
        FROM transport_sheet_changes WHERE sheet_id=$1 ORDER BY created_at DESC""", sheet_id)
    refuels = await pool.fetch(                                     # v56
        "SELECT id,liters,refuel_at FROM transport_sheet_refuels WHERE sheet_id=$1 ORDER BY refuel_at,id",
        sheet_id)
    result = _sheet_dict(row)
    result["changes"] = [dict(c) for c in changes]
    result["refuels"] = [dict(x) for x in refuels]
    return result
